from dnac import DNAC
import os
import openpyxl

if __name__ == '__main__':
    try:
        USERNAME = os.environ['DNAC_USERNAME']
        PASSWORD = os.environ['DNAC_PASSWORD']
        DNAC_IP = os.environ['DNAC_IP']
    except KeyError:
        print('Switch ENV Variables Not Found. Exiting...')
        exit(1)

    dnac = DNAC(DNAC_IP, USERNAME, PASSWORD)
    auth_token = "eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCJ9.eyJzdWIiOiI2NjI2YmI1MTNhZjI2YTY0Mzk1MmZmYzEiLCJhdXRoU291cmNlIjoiZXh0ZXJuYWwiLCJ0ZW5hbnROYW1lIjoiVE5UMCIsInJvbGVzIjpbIjY1ZmNhZGI2OGM3NWUwNDAwMzY2YzdhMCJdLCJ0ZW5hbnRJZCI6IjY1ZmNhZGI1OGM3NWUwNDAwMzY2Yzc5ZSIsImV4cCI6MTcxODUyNjI1MCwiaWF0IjoxNzE4NTIyNjUwLCJqdGkiOiJkOTUwNGI1Mi0yYzBmLTQwODctOTEzZS0zMTMyY2FjYTBlZmMiLCJ1c2VybmFtZSI6InZvbG9keW15cmsxIn0.nIOuWYLJMjMQ8cLtODtxluY_FVGMZTaNK6PHniRAH0ReJ4cztJXRwxQFv2rJahAQMho4OHgiq5Qv0K_8aSdyFlr82QrxEMZQpnLWS8w0VphwB9HdIQAFdjOhaASwyJSJenfSOj5-fXDGH-YGfTuR762y2ZlHP-R3rQtoD0-HaKEYJyZQJluYH1OAl0DHz_bsxIuHrMGQonWJdMXER2iXYy1MLIyKH4Or9N5Z3FUGEPGAZUtrnFEf9ag21KO62FNlRkpCdMI4wWyS8EYFy9jNnl_KTbD_wf7de3NUwnpUHMxAv4RUfw4LnZlU0-7Ogb_PZX1wUTV9bEOWaAm8TTBzjQ"
    dnac.session.headers.update({'X-Auth-Token': auth_token,
                                 'Content-Type': 'application/json', 'Accept': 'application/json'})
#    dnac.auth()

    # working with:
    # fabric_site_name = '' # SDA Fabric Site Name. String.
    # my_switch_ip = '' # Switch IP Address.

    # no github publishing ;)
    fabric_site_name = os.environ['fabric_site_name']
    my_switch_ip = os.environ['my_switch_ip']

    # define ports to be added.
    # DNAC uses VLAN names instead of VLAN ID to assign port
    # later, for easier, manipulation we can list of the VLANs, crete dict, and get proper name using ID.
    # for now I will use known names

    # example for AP, ACCESS, and Trunk ports:
    ports = [
        {'interfaceName': 'GigabitEthernet1/0/10', 'interfaceDescription': 'TEST1',
         'connectedDeviceType': 'USER_DEVICE', 'dataVlanName': 'PC_VLAN', 'voiceVlanName': 'VOICE_VLAN'},
        {'interfaceName': 'GigabitEthernet1/0/11', 'interfaceDescription': 'TEST2',
         'connectedDeviceType': 'ACCESS_POINT', 'dataVlanName': 'AP_VLAN_2200', 'voiceVlanName': ''},
        {'interfaceName': 'GigabitEthernet1/0/12', 'interfaceDescription': 'TEST3',
         'connectedDeviceType': 'TRUNKING_DEVICE', 'dataVlanName': '', 'voiceVlanName': ''},
    ]

    # get DNAC sites
    dnac_sites = dnac.get_sites()

    # get fabric ID using site ID
    my_site_id = dnac_sites[fabric_site_name]['id']
    my_site_hie = dnac_sites[fabric_site_name]['groupNameHierarchy']
    my_fabric_id = dnac.get_fabric_id(my_site_id)

    # get device ID for B535 SW:
    my_switch_id = dnac.get_device_id(my_switch_ip)

    # remove if assigned. One by one. Each task should be completed before submitting new one.
    # for port in ports:
    #    interface_name = port['interfaceName']
    #    if dnac.is_port_assigned(my_switch_ip, interface_name):
    #        ret = dnac.delete_port_assignment(my_fabric_id, my_switch_id, interface_name)
    #        taskId = ret['taskId']
    #        print(f'Port Deletion task {taskId} was submitted. Waiting for execution')
    #        dnac.wait_for_task(taskId)

    # wait two minutes, to test id, I want to go to DNAC and SW and see that ports were removed
    # print('wait 120 seconds')
    # time.sleep(120)

    # add ports. Bulk operations. Single Task.
    # ret = dnac.assign_ports(my_fabric_id, my_switch_id, ports)
    # taskId = ret['taskId']
    # print(f'Port assignment task {taskId} was submitted. Waiting for execution')
    # dnac.wait_for_task(taskId)

    # Add test AnycastGateway for 10.6.23.0/25 (10.6.23.1)
    add_subnet = '10.60.100.0/24'
    add_subnet_gw = '10.60.100.1'

    ip_pool_name = '10.60.100.0_24'
    VN_ID = 'IOT_VN'
    VLAN_NAME = 'IOT_TEST'
    VLAN_ID = '987'
    DHCP = ['10.6.14.10', '10.16.171.10']
    DNS = ['10.6.14.10', '10.5.14.10']


    #delete test
    #if dnac.get_anycast_gateway(my_fabric_id, ip_pool_name):
    #   dnac.delete_anycast_gateway(my_site_hie, VN_ID, ip_pool_name)
    #

    # if subnet does not exist, create new. If exists, get name, id and delete.
    #if not dnac.is_subnet_exit(my_site_id, add_subnet):
    #    parent_subnet = dnac.get_subnet_global_parent(add_subnet)
    #    if not parent_subnet:
    #        print(f'Error: {add_subnet} does not have parent Global Subnet. Please fix and try again')
    #    else:
    #        print(f'Reserving pool: {add_subnet}')
    #        dnac.reserve_subnet(my_site_id, parent_subnet, add_subnet, add_subnet_gw, ip_pool_name, DNS, DHCP)
    #else:
    #    ip_pool_name = dnac.get_ippool_name(my_site_id, add_subnet)
    #    ip_pool_id = dnac.get_ippool_id(my_site_id, add_subnet)
    #    print(f'{ip_pool_name} with ID:{ip_pool_id} exists. Removing' )
    #    dnac.release_subnet(ip_pool_id)


    # if Anycast GW does not exit, create new.
    #if not dnac.get_anycast_gateway(my_fabric_id, ip_pool_name):
    #    dnac.add_anycast_gateway(my_site_hie, VN_ID, ip_pool_name, VLAN_NAME, VLAN_ID)
    #    pool = dnac.get_anycast_gateway(my_fabric_id, ip_pool_name)
    # else:
        ## N/A dnac.update_anycast_gateway(my_site_hie, VN_ID, ip_pool_name, VLAN_NAME)
        # pass


#### Sat Jun 15

    # read xcel
    wb_obj = openpyxl.load_workbook('./all_b535_anycast_gw_and_l2.xlsx')
    sheet_obj = wb_obj['AnycastGateways']

    max_row = sheet_obj.max_row
    anycast_gw_list = []

    for row in range(2, max_row + 1):
        anycast_gw = {}
        anycast_gw['vlan_id'] = sheet_obj.cell(row=row, column=1).value
        segment_type = sheet_obj.cell(row=row, column=2).value
        anycast_gw['vlan_name'] = sheet_obj.cell(row=row, column=3).value
        anycast_gw['segment_type'] = segment_type
        if segment_type == 'Layer3':
            anycast_gw['subnet'] = sheet_obj.cell(row=row, column=5).value + sheet_obj.cell(row=row, column=6).value
            anycast_gw['gateway_ip'] = sheet_obj.cell(row=row, column=7).value
            anycast_gw['vn_name'] = sheet_obj.cell(row=row, column=8).value
            anycast_gw['pool_name'] = sheet_obj.cell(row=row, column=4).value
        else:
            anycast_gw['pool_name'] = sheet_obj.cell(row=row, column=3).value
            anycast_gw['vn_name'] = 'USER_VN'
        anycast_gw_list.append(anycast_gw)

#    print(anycast_gw_list)

    for gw in anycast_gw_list:
        VN_ID = gw['vn_name']
        ip_pool_name = gw['pool_name']
        VLAN_NAME = gw['vlan_name'][0:32]
        VLAN_ID = str(gw['vlan_id'])
        if gw['segment_type'] == 'Layer3':
            add_subnet = gw['subnet']
            add_subnet_gw = gw['gateway_ip']
        else:
            add_subnet = ''
            add_subnet_gw = ''
        # print(VLAN_ID, VLAN_NAME, ip_pool_name, VN_ID, add_subnet, add_subnet_gw)


    #exit(1)
        # reserve L3 pools
        if gw['segment_type'] == 'Layer3':
            if not dnac.is_subnet_exit(my_site_id, add_subnet):
                parent_subnet = dnac.get_subnet_global_parent(add_subnet)
                if not parent_subnet:
                    print(f'Error: {add_subnet} does not have parent Global Subnet. Please fix and try again')
                else:
                    print(f'Reserving pool: {add_subnet}')
                    dnac.reserve_subnet(my_site_id, parent_subnet, add_subnet, add_subnet_gw, ip_pool_name, DNS, DHCP)
            else:
                ip_pool_id = 'N/A'
                ip_pool_name = dnac.get_ippool_name(my_site_id, add_subnet)
                #ip_pool_id = dnac.get_ippool_id(my_site_id, add_subnet)
                print(f'{ip_pool_name} for subnet {add_subnet} exists with ID:{ip_pool_id}' )

            # if Anycast GW does not exit, create new.
            if not dnac.get_anycast_gateway(my_fabric_id, ip_pool_name):
                dnac.add_anycast_gateway(my_site_hie, VN_ID, ip_pool_name, VLAN_NAME, VLAN_ID)

            # delete GW if rollback is required.
            # if dnac.get_anycast_gateway(my_fabric_id, ip_pool_name):
            #    pool = dnac.get_anycast_gateway(my_fabric_id, ip_pool_name)

        #if gw['segment_type'] == 'Layer2':
        #    dnac.add_l2segment(my_site_hie, VLAN_NAME, VLAN_ID)


## ToDo: truncate VLAN name inside of add anycast gw and add l2 segment.
## ToDo: add logger to debug.
## ToDo: add 429 Error handling.